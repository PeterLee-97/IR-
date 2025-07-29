#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
IR 보고서 모니터링 & 요약 메일 발송 스크립트
- requests 403/동적 페이지: chrome-headless-shell + Selenium fallback
- selector가 잘못되면 자동 판별 후 fallback으로 PDF 링크 탐색
- Nintendo: headless → “Earnings Release” 텍스트 + PDF(연도 무관) 최신 2개
- Take-Two Interactive: headless만, “Earnings Release” 텍스트 최신 2개
- Apple, EA, Sony, Netmarble, Roblox: headless만
- KakaoGames: headless만, 클릭 다운로드
- NCSoft, ShiftUp, NetEase, Kingnet: 각각 특이 로직
- AI 요약: Incognito 모드로 매번 새 창을 띄워, 최대 360초(6분) 대기 후 “이상 문서 끝.” 체크, 2회 재시도
"""

import os
import re
import time
import logging
import threading
import smtplib
from pathlib import Path
from urllib.parse import urljoin
from datetime import datetime

import requests
from bs4 import BeautifulSoup
import openpyxl

from email.mime.text import MIMEText
from email.header import Header

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import (
    TimeoutException,
    InvalidSelectorException,
    StaleElementReferenceException,
)
from selenium.webdriver.chrome.service import Service as ChromeService

# ───────────────────────────────────────────────────────────────────────────────
# 0. 설정
# ───────────────────────────────────────────────────────────────────────────────

COMPANIES_EXCEL_PATH      = 'companies.xlsx'
FOUND_REPORTS_EXCEL_PATH  = 'found_reports.xlsx'
CHECK_INTERVAL_SECONDS    = 600

LOCAL_PDF_PATH_PREFIX     = "downloaded_report"
MAX_CONCURRENT_BROWSERS   = 1
browser_semaphore         = threading.Semaphore(MAX_CONCURRENT_BROWSERS)
HEADLESS_CHROME_BIN       = "/path/to/chrome-headless-shell"
CICI_CHAT_URL             = "https://www.cici.com/chat/"
SMTP_SERVER               = "smtp.gmail.com"
SMTP_PORT                 = 587
SMTP_USER                 = os.environ.get('SMTP_GMAIL_USER', 'rbd.autobot@gmail.com')
SMTP_PASS                 = os.environ.get('SMTP_GMAIL_PASS', 'bzdm voqt oxtg ekyv')
RECIPIENTS                = ["chonghyok@nexon.co.kr",]

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(threadName)s - %(levelname)s - %(message)s'
)

PDF_REGEX = re.compile(r'\.pdf(\?.*)?$', re.IGNORECASE)
NINTENDO_YEAR_REGEX = re.compile(r'/ir/pdf/(20(?:2[5-9]|[3-9]\d))/')

# ───────────────────────────────────────────────────────────────────────────────
# 1. Selenium 드라이버 (incognito 옵션 추가)
# ───────────────────────────────────────────────────────────────────────────────

def build_chrome_driver(headless: bool = True, incognito: bool = False, download_dir: str = None, page_load_timeout: int = 60) -> webdriver.Chrome:
    options = webdriver.ChromeOptions()
    if headless:
        options.add_argument("--headless=new")
    if incognito:
        options.add_argument("--incognito")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--window-size=1920x1080")
    options.add_argument("--disable-gpu")
    options.add_argument("--disable-blink-features=AutomationControlled")
    options.add_argument("--ignore-certificate-errors")
    options.add_argument(
        "user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/126.0.0.0 Safari/537.36"
    )
    options.add_experimental_option("excludeSwitches", ["enable-automation"])
    options.add_experimental_option('useAutomationExtension', False)
    if download_dir:
        prefs = {
            "download.default_directory": download_dir,
            "download.prompt_for_download": False,
            "plugins.always_open_pdf_externally": True
        }
        options.add_experimental_option("prefs", prefs)
    options.binary_location = HEADLESS_CHROME_BIN

    driver = webdriver.Chrome(service=ChromeService(), options=options)
    driver.set_page_load_timeout(page_load_timeout)
    return driver

# ───────────────────────────────────────────────────────────────────────────────
# 2. 엑셀 I/O
# ───────────────────────────────────────────────────────────────────────────────

def load_companies_from_excel(filepath):
    wb = openpyxl.load_workbook(filepath)
    sh = wb.active
    companies = [
        {'name': row[0], 'url': row[1], 'selector': row[2], 'prompt': row[3]}
        for row in sh.iter_rows(min_row=2, values_only=True) if row and all(row[:4])
    ]
    logging.info(f"'{filepath}'에서 {len(companies)}개 기업 정보를 불러왔습니다.")
    return companies

def load_found_urls_from_excel(filepath):
    if not Path(filepath).is_file():
        wb = openpyxl.Workbook()
        wb.active.append(["Report URL","Timestamp"])
        wb.save(filepath)
        return set()
    wb = openpyxl.load_workbook(filepath)
    return {row[0] for row in wb.active.iter_rows(min_row=2, values_only=True) if row and row[0]}

def save_url_to_excel(filepath, url):
    wb = openpyxl.load_workbook(filepath)
    wb.active.append([url, time.strftime('%Y-%m-%d %H:%M:%S')])
    wb.save(filepath)

# ───────────────────────────────────────────────────────────────────────────────
# 3. HTML 로딩 & PDF 링크 파싱
# ───────────────────────────────────────────────────────────────────────────────

def get_html_by_requests(url):
    resp = requests.get(
        url,
        headers={'User-Agent':'Mozilla/5.0','Accept-Language':'en-US,en;q=0.9,ko;q=0.8'},
        timeout=20
    )
    resp.raise_for_status()
    return resp.text

def get_html_by_headless(url, wait_selector=None, timeout=20):
    driver = build_chrome_driver(headless=True, page_load_timeout=timeout)
    try:
        driver.get(url)
        if wait_selector:
            try:
                WebDriverWait(driver, timeout).until(
                    EC.presence_of_all_elements_located((By.CSS_SELECTOR, wait_selector))
                )
            except (TimeoutException, InvalidSelectorException):
                pass
        return driver.page_source
    finally:
        driver.quit()

def safe_select(soup, selector):
    try:
        return soup.select(selector)
    except:
        return []

def extract_links(base, tags):
    return list({
        urljoin(base, t.get('href'))
        for t in tags
        if t.get('href') and PDF_REGEX.search(t.get('href'))
    })

def fallback_all_pdf(soup, base):
    return list({
        urljoin(base, a['href'])
        for a in soup.find_all('a', href=True)
        if PDF_REGEX.search(a['href'])
    })

def scrape_reports_for_company(company):
    name, url, sel = company['name'], company['url'], company['selector']
    logging.info(f"[{name}] 사이트 확인: {url}")

    # Nintendo
    if name.lower() == 'nintendo':
        html = get_html_by_headless(url, wait_selector='.corp_ir_2018-newsItem__text')
        soup = BeautifulSoup(html, 'html.parser')
        tags = [
            t for t in soup.select('.corp_ir_2018-newsItem__text.pdfinfo a.corp_ir_2018-newsLink--pdf')
            if 'Earnings Release' in t.get_text()
        ]
        links = extract_links(url, tags)
        links = sorted(links, reverse=True)[:2]
        logging.info(f"[{name}] 링크 발견: {len(links)}개")
        return links

    # Take-Two Interactive: 최신 2개 Earnings Release
    if 'take-two' in name.lower():
        html = get_html_by_headless(url, wait_selector='.css-1s1i6cp')
        soup = BeautifulSoup(html, 'html.parser')
        records = []
        for block in soup.select('.css-1s1i6cp'):
            date_txt = (block.select_one('p.css-19b5k1g') or BeautifulSoup('', 'html.parser')) \
                       .get_text().split(' at')[0]
            try:
                dt = datetime.strptime(date_txt, '%B %d, %Y')
            except:
                continue
            for a in block.select('a.css-1wvogm1'):
                if 'Earnings Release' in a.get_text() and a.get('href'):
                    records.append((dt, urljoin(url, a['href'])))
        links = [link for _, link in sorted(records, key=lambda x: x[0], reverse=True)[:2]]
        logging.info(f"[{name}] 링크 발견: {len(links)}개")
        return links

    # KakaoGames
    if name.lower().startswith(('카카오','kakao')):
        html = get_html_by_headless(url, wait_selector=company['selector'])
        soup = BeautifulSoup(html, 'html.parser')
        tags = soup.select(f"{company['selector']}[data-earningreporturl$='.pdf']")
        links = [
            urljoin(url, t['data-earningreporturl'])
            for t in tags
            if t.get('data-earningreporturl') and PDF_REGEX.search(t['data-earningreporturl'])
        ]
        logging.info(f"[{name}] 링크 발견: {len(links)}개")
        return links

    # Apple, EA, Sony, Netmarble, Roblox
    if name.lower() in (
        'apple','electronic arts','ea','Sony',
        'netmarble','넷마블','roblox'
    ):
        html = get_html_by_headless(url, wait_selector=sel)
        soup = BeautifulSoup(html, 'html.parser')
        tags = safe_select(soup, sel)
        links = extract_links(url, tags) or fallback_all_pdf(soup, url)
        logging.info(f"[{name}] 링크 발견: {len(links)}개")
        return links

    # NCSoft
    if name.lower() in ('ncsoft', '엔씨소프트'):
        html = get_html_by_headless(url, wait_selector='td.g_text2')
        soup = BeautifulSoup(html, 'html.parser')
        tags = soup.select("td.g_text2 a.g_btn[href*='fileDownload']")
        records = []
        for a in tags:
            href = a['href']
            m = re.search(r"fileDownload\(\s*'([^']+)'\s*,\s*'([^']+\.pdf)'\s*\)", href)
            if not m:
                continue
            base_url, raw_fname = m.group(1), m.group(2).strip()
            date_str = base_url.rstrip('/').split('/')[-1]
            try:
                dt = datetime.strptime(date_str, '%Y%m%d')
            except ValueError:
                dt = datetime.min
            records.append((dt, f"{base_url}/{raw_fname}"))
        records.sort(key=lambda x: x[0], reverse=True)
        links = [link for _, link in records[:2]]
        logging.info(f"[{name}] 링크 발견: {len(links)}개")
        return links

    # ShiftUp
    if name.lower().startswith(('shiftup','시프트업')):
        html = get_html_by_headless(url, wait_selector='a.downloadBtn')
        soup = BeautifulSoup(html, 'html.parser')
        tags = soup.select('a.downloadBtn')
        links = [urljoin(url, a['href']) for a in tags][:2]
        logging.info(f"[{name}] 링크 발견: {len(links)}개")
        return links

    # NetEase
    if 'netease' in name.lower() or name.startswith('넷이즈'):
        html = get_html_by_headless(url, wait_selector='div.file--application-pdf')
        soup = BeautifulSoup(html, 'html.parser')
        tags = soup.select('div.file--application-pdf a[type="application/pdf"]')
        records = []
        for a in tags:
            m = re.search(r'/(\d{4})/(\d{2})/(\d{2})/', a['href'])
            if m:
                dt = datetime(*map(int, m.groups()))
            else:
                dt = datetime.min
            records.append((dt, urljoin(url, a['href'])))
        links = [link for _, link in sorted(records, key=lambda x: x[0], reverse=True)[:2]]
        logging.info(f"[{name}] 링크 발견: {len(links)}개")
        return links

    # Kingnet
    if name.lower().startswith(('kingnet','킹넷')):
        driver = build_chrome_driver(headless=True)
        try:
            driver.get(url)
            elem = WebDriverWait(driver, 20).until(
                EC.element_to_be_clickable((By.CSS_SELECTOR, '여기에-셀렉터'))
            )
            elem.click()
            WebDriverWait(driver, 10).until(lambda d: len(d.window_handles) > 1)
            main = driver.current_window_handle
            for h in driver.window_handles:
                if h != main:
                    driver.switch_to.window(h)
                    break
            links = [driver.current_url]
        finally:
            driver.quit()
        logging.info(f"[{name}] 링크 발견: {len(links)}개")
        return links

    # 기타
    html = None
    try:
        cont = get_html_by_requests(url)
        if len(cont) > 500:
            html = cont
    except:
        pass
    if not html:
        html = get_html_by_headless(url, wait_selector=sel)
    soup = BeautifulSoup(html, 'html.parser')
    tags = safe_select(soup, sel)
    links = extract_links(url, tags) or fallback_all_pdf(soup, url)
    logging.info(f"[{name}] 링크 발견: {len(links)}개")
    return links

# ───────────────────────────────────────────────────────────────────────────────
# 4. 보고서 처리 & 메일 (Worker)
# ───────────────────────────────────────────────────────────────────────────────

def process_and_email_report(pdf_url, company_info, lock, shared_urls_set):
    name = company_info['name']
    browser_semaphore.acquire()
    logging.info(f"[{name}] 처리 시작! (남은 브라우저 슬롯: {browser_semaphore._value})")

    thread_id      = threading.get_ident()
    local_pdf_path = f"{LOCAL_PDF_PATH_PREFIX}_{thread_id}.pdf"
    success        = False

    try:
        # 1) PDF 다운로드
        logging.info(f"[{name}] PDF 다운로드 → {pdf_url}")
        resp = requests.get(pdf_url, headers={'User-Agent': 'Mozilla/5.0'}, timeout=60, stream=True)
        resp.raise_for_status()
        with open(local_pdf_path, "wb") as f:
            for chunk in resp.iter_content(chunk_size=256*1024):
                if chunk:
                    f.write(chunk)
        logging.info(f"[{name}] PDF 저장 완료: {local_pdf_path}")

        # 2) AI 요약: 최대 2회 재시도, 각 시도마다 최대 360초 대기 & 10초마다 '이상 문서 끝.' 체크
        summary_html = None
        email_subject = None

        for attempt in range(10):
            driver = build_chrome_driver(headless=False, incognito=True, page_load_timeout=60)
            wait   = WebDriverWait(driver, 60)
            try:
                driver.get(CICI_CHAT_URL)
                wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, "div[data-testid='upload_file_button']"))).click()
                wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, "input[type='file']"))).send_keys(os.path.abspath(local_pdf_path))

                chat = wait.until(EC.element_to_be_clickable((By.XPATH, "//textarea[@data-testid='chat_input_input']")))
                chat.clear()
                chat.send_keys(company_info['prompt'])

                for _ in range(3):
                    try:
                        btn = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, "button[data-testid='chat_input_send_button']")))
                        driver.execute_script("arguments[0].click();", btn)
                        break
                    except StaleElementReferenceException:
                        time.sleep(1)
                else:
                    raise Exception("전송 버튼 클릭 실패")

                # 두 번째 메시지 위치 대기
                second_elem = WebDriverWait(driver, 60).until(
                    EC.presence_of_element_located((By.XPATH, "(//div[@data-testid='message_text_content'])[2]"))
                )

                # 최대 360초 동안 10초마다 '이상 문서 끝.' 체크
                timeout = 180
                interval = 10
                elapsed = 0
                text_only = ""
                while elapsed < timeout:
                    text_only = second_elem.text.strip()
                    if "이상 문서 끝." in text_only:
                        break
                    time.sleep(interval)
                    elapsed += interval

                if len(text_only) <= 50:
                    logging.warning(f"[{name}] 요약 본문 짧음({len(text_only)}글자)→ 재시도 {attempt+1}/2")
                    driver.quit()
                    continue

                summary_html = second_elem.get_attribute("innerHTML") or ""
                try:
                    title_elem = second_elem.find_element(By.XPATH, "./div[1]")
                    email_subject = title_elem.get_attribute("textContent").strip() or f"{name} 재무 보고서 요약"
                except:
                    email_subject = f"{name} 재무 보고서 요약 (제목 추출 실패)"

                logging.info(f"[{name}] AI 응답 완료. 제목: {email_subject}")
                driver.quit()
                break

            except Exception as e:
                driver.quit()
                if attempt == 1:
                    raise Exception(f"AI 요약 실패: {e}")
                else:
                    continue

        if not summary_html:
            raise Exception("AI 요약 본문 획득 실패")

        # 3) 메일 전송
        msg = MIMEText(
            f"<p><b>기업:</b> {name}<br>"
            f"<b>원본 보고서 링크:</b> <a href=\"{pdf_url}\">{pdf_url}</a></p><hr>"
            f"{summary_html}",
            "html", "utf-8"
        )
        msg["Subject"] = Header(email_subject, "utf-8")
        msg["From"]    = SMTP_USER
        msg["To"]      = ", ".join(RECIPIENTS)

        with smtplib.SMTP(SMTP_SERVER, SMTP_PORT) as smtp:
            smtp.starttls()
            smtp.login(SMTP_USER, SMTP_PASS)
            smtp.sendmail(SMTP_USER, RECIPIENTS, msg.as_string())
        logging.info(f"[{name}] 이메일 전송 성공 → {', '.join(RECIPIENTS)}")
        success = True

    except Exception as e:
        logging.error(f"[{name}] 보고서 처리 오류: {e}", exc_info=True)

    finally:
        if os.path.exists(local_pdf_path):
            try: os.remove(local_pdf_path)
            except: pass
        with lock:
            if success:
                save_url_to_excel(FOUND_REPORTS_EXCEL_PATH, pdf_url)
            else:
                shared_urls_set.discard(pdf_url)
        browser_semaphore.release()
        logging.info(f"[{name}] 처리 완료. 남은 브라우저 슬롯: {browser_semaphore._value}")

# ───────────────────────────────────────────────────────────────────────────────
# 5. 메인
# ───────────────────────────────────────────────────────────────────────────────

def main():
    if not (SMTP_USER and SMTP_PASS):
        logging.error("SMTP 인증 정보가 없습니다. 종료합니다.")
        return

    logging.info("IR 자동화 시작")
    companies = load_companies_from_excel(COMPANIES_EXCEL_PATH)
    seen      = load_found_urls_from_excel(FOUND_REPORTS_EXCEL_PATH)
    lock      = threading.Lock()

    while True:
        new_reports = []
        for comp in companies:
            try:
                for link in scrape_reports_for_company(comp):
                    if link not in seen:
                        seen.add(link)
                        new_reports.append((link, comp))
                        logging.info(f"[!!!] 신규: [{comp['name']}] {link}")
            except Exception as e:
                logging.error(f"[{comp['name']}] 오류: {e}", exc_info=True)
            time.sleep(2)

        if new_reports:
            logging.info(f"{len(new_reports)}건 처리 시작")
            for link, comp in new_reports:
                t = threading.Thread(
                    target=process_and_email_report,
                    args=(link, comp, lock, seen),
                    name=f"Worker-{comp['name']}"
                )
                t.start()
        else:
            logging.info("신규 보고서 없음")

        time.sleep(CHECK_INTERVAL_SECONDS)

if __name__ == "__main__":
    main()
